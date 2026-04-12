"""Generate Step_4.0_Scraping_Progress_Report.md and .xlsx from master DB."""

import sqlite3
import os
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_DB = os.path.join(SCRIPT_DIR, "Data", "Step_5_master_reviews.db")
MD_PATH = os.path.join(SCRIPT_DIR, "Data", "Step_4.0_Scraping_Progress_Report.md")
XLSX_PATH = os.path.join(SCRIPT_DIR, "Data", "Step_4.0_Scraping_Progress_Report.xlsx")

conn = sqlite3.connect(MASTER_DB)
cur = conn.cursor()

today = datetime.now().strftime("%Y-%m-%d")

# ── Overall stats ────────────────────────────────────────────────────────────
cur.execute("SELECT COUNT(*) FROM stores")
total_stores = cur.fetchone()[0]

cur.execute("SELECT COUNT(*) FROM reviews")
total_reviews = cur.fetchone()[0]

cur.execute("SELECT COUNT(*) FROM stores WHERE target_reviews IS NULL OR target_reviews = 0")
no_target = cur.fetchone()[0]

cur.execute("""
    SELECT scrape_status, COUNT(*),
           COALESCE(SUM(target_reviews),0),
           COALESCE(SUM(reviews_scraped),0)
    FROM stores GROUP BY scrape_status
""")
status_rows = {r[0]: {"count": r[1], "target": r[2], "scraped": r[3]} for r in cur.fetchall()}

complete = status_rows.get("complete", {"count": 0, "target": 0, "scraped": 0})
incomplete = status_rows.get("incomplete", {"count": 0, "target": 0, "scraped": 0})
pending = status_rows.get("pending", {"count": 0, "target": 0, "scraped": 0})

total_target = complete["target"] + incomplete["target"] + pending["target"]
total_scraped = complete["scraped"] + incomplete["scraped"] + pending["scraped"]
total_remaining = max(0, incomplete["target"] - incomplete["scraped"]) + pending["target"]

# ── By category ──────────────────────────────────────────────────────────────
cur.execute("""
    SELECT
        COALESCE(NULLIF(category,''), NULLIF(type,''), 'Uncategorized') as cat,
        COUNT(*) as cnt,
        SUM(CASE WHEN scrape_status='complete' THEN 1 ELSE 0 END) as done,
        SUM(CASE WHEN scrape_status='incomplete' THEN 1 ELSE 0 END) as inc,
        SUM(CASE WHEN scrape_status='pending' THEN 1 ELSE 0 END) as pend,
        COALESCE(SUM(reviews_scraped),0) as collected,
        COALESCE(SUM(target_reviews),0) as target_sum,
        COALESCE(SUM(CASE WHEN scrape_status != 'complete' AND target_reviews > 0
                      THEN MAX(0, target_reviews - reviews_scraped) ELSE 0 END),0) as remain
    FROM stores
    GROUP BY cat
    ORDER BY target_sum DESC
""")
cat_rows = cur.fetchall()

# Split into major (target > 0) and minor
major_cats = [r for r in cat_rows if r[6] > 0]
minor_cats = [r for r in cat_rows if r[6] == 0]

# ── Flagged mismatches ───────────────────────────────────────────────────────
cur.execute("""
    SELECT store_id, store_name, place_id, full_address, target_reviews,
           reviews_scraped, (reviews_scraped - target_reviews) as diff,
           COALESCE(NULLIF(category,''), NULLIF(type,''), 'Uncategorized') as cat
    FROM stores
    WHERE reviews_scraped > target_reviews + 5 AND target_reviews > 0
    ORDER BY diff DESC
""")
flagged = cur.fetchall()

# ── Incomplete by category ───────────────────────────────────────────────────
cur.execute("""
    SELECT
        COALESCE(NULLIF(category,''), NULLIF(type,''), 'Uncategorized') as cat,
        COUNT(*) as cnt,
        SUM(MAX(0, target_reviews - reviews_scraped)) as remain
    FROM stores
    WHERE scrape_status = 'incomplete' AND target_reviews > 0
    GROUP BY cat ORDER BY remain DESC
""")
inc_cats = cur.fetchall()

# ── Pending by category ─────────────────────────────────────────────────────
cur.execute("""
    SELECT
        COALESCE(NULLIF(category,''), NULLIF(type,''), 'Uncategorized') as cat,
        COUNT(*) as cnt,
        COALESCE(SUM(target_reviews),0) as target
    FROM stores
    WHERE scrape_status = 'pending' AND target_reviews > 0
    GROUP BY cat ORDER BY target DESC
""")
pend_cats = cur.fetchall()

# ── Batch summary ────────────────────────────────────────────────────────────
batch_info = []
try:
    import glob as gl
    for bp in sorted(gl.glob(os.path.join(SCRIPT_DIR, "Data", "batch_*.db"))):
        bconn = sqlite3.connect(bp)
        bcur = bconn.cursor()
        bcur.execute("SELECT COUNT(*), SUM(target_reviews), SUM(master_reviews), SUM(target_reviews - master_reviews) FROM stores")
        r = bcur.fetchone()
        bcur.execute("SELECT MIN(target_reviews), MAX(target_reviews) FROM stores")
        mm = bcur.fetchone()
        batch_info.append({
            "name": os.path.basename(bp),
            "stores": r[0], "target": r[1], "already": r[2], "delta": r[3],
            "lo": mm[0], "hi": mm[1],
        })
        bconn.close()
except Exception:
    pass

conn.close()

# ══════════════════════════════════════════════════════════════════════════════
# Generate Markdown
# ══════════════════════════════════════════════════════════════════════════════
pct_complete = round(total_scraped / total_target * 100, 1) if total_target else 0
lines = []
a = lines.append

a(f"# Scraping Progress Report - {today}")
a("")
a("## Overall Summary")
a("")
a("| Metric | Value |")
a("|--------|-------|")
a(f"| Total stores | {total_stores:,} |")
a(f"| Stores complete | {complete['count']:,} ({round(complete['count']/total_stores*100,1)}%) |")
a(f"| Stores incomplete | {incomplete['count']:,} ({round(incomplete['count']/total_stores*100,1)}%) |")
a(f"| Stores pending | {pending['count']:,} ({round(pending['count']/total_stores*100,1)}%) |")
a(f"| Stores with no target | {no_target} |")
a(f"| Reviews collected | {total_scraped:,} |")
a(f"| Reviews in DB (incl. overshoot) | {total_reviews:,} |")
a(f"| Target total (capped @1500) | {total_target:,} |")
a(f"| Reviews remaining | {total_remaining:,} |")
a(f"| Completion % | {pct_complete}% |")
a("")

a("## By Category")
a("")
a("| Category | Stores | Complete | Incomplete | Pending | Collected | Target | Remaining | % Done |")
a("|----------|--------|----------|------------|---------|-----------|--------|-----------|--------|")
t_cnt = t_c = t_i = t_p = t_col = t_tar = t_rem = 0
for r in major_cats:
    cat, cnt, done, inc, pend, collected, target, remain = r
    pct = round(collected / target * 100, 1) if target else 0
    a(f"| {cat} | {cnt:,} | {done:,} | {inc:,} | {pend:,} | {collected:,} | {target:,} | {remain:,} | {pct}% |")
    t_cnt += cnt; t_c += done; t_i += inc; t_p += pend; t_col += collected; t_tar += target; t_rem += remain
pct_tot = round(t_col / t_tar * 100, 1) if t_tar else 0
a(f"| **TOTAL** | **{t_cnt:,}** | **{t_c:,}** | **{t_i:,}** | **{t_p:,}** | **{t_col:,}** | **{t_tar:,}** | **{t_rem:,}** | **{pct_tot}%** |")
a("")

if minor_cats:
    a(f"*{len(minor_cats)} categories with no target reviews ({sum(r[1] for r in minor_cats)} stores) excluded from table above.*")
    a("")

a("## By Status")
a("")
a("| Status | Stores | Target | Collected | Remaining | % Done |")
a("|--------|--------|--------|-----------|-----------|--------|")
for name, d in [("complete", complete), ("incomplete", incomplete), ("pending", pending)]:
    rem = max(0, d["target"] - d["scraped"])
    pct = round(d["scraped"] / d["target"] * 100, 1) if d["target"] else 0
    a(f"| {name} | {d['count']:,} | {d['target']:,} | {d['scraped']:,} | {rem:,} | {pct}% |")
a("")

a("## Work Remaining by Priority")
a("")
a("### Incomplete stores (have some reviews, need more)")
a("")
a("| Category | Stores | Reviews Remaining |")
a("|----------|--------|-------------------|")
for r in inc_cats:
    a(f"| {r[0]} | {r[1]:,} | {r[2]:,} |")
a("")

a("### Pending stores (not yet started)")
a("")
a("| Category | Stores | Target |")
a("|----------|--------|--------|")
for r in pend_cats:
    a(f"| {r[0]} | {r[1]:,} | {r[2]:,} |")
a("")

if batch_info:
    a("## Batch Files for Linux Scraper")
    a("")
    a(f"*{sum(b['stores'] for b in batch_info):,} stores across {len(batch_info)} batches, {sum(b['delta'] for b in batch_info):,} reviews to scrape*")
    a("")
    a("| Batch | Stores | Target Range | Already Have | To Scrape |")
    a("|-------|--------|-------------|--------------|-----------|")
    for b in batch_info:
        a(f"| {b['name']} | {b['stores']:,} | {b['lo']:,}-{b['hi']:,} | {b['already']:,} | {b['delta']:,} |")
    a("")

if flagged:
    a("## Flagged: Possible Wrong-Store Scrapes")
    a("")
    a(f"*{len(flagged)} stores where reviews_scraped exceeds target by >5 — may have scraped a different location with the same name.*")
    a("")
    a("| Store ID | Name | Target | Scraped | Excess | Category |")
    a("|----------|------|--------|---------|--------|----------|")
    for r in flagged:
        a(f"| {r[0]} | {r[1]} | {r[4]:,} | {r[5]:,} | +{r[6]:,} | {r[7]} |")
    a("")

with open(MD_PATH, "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")
print(f"Wrote {MD_PATH}")

# ══════════════════════════════════════════════════════════════════════════════
# Generate Excel
# ══════════════════════════════════════════════════════════════════════════════
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed, installing...")
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
pct_fmt = "0.0%"
num_fmt = "#,##0"


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def style_row(ws, row, ncols, is_total=False):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        if is_total:
            cell.font = total_font
            cell.fill = total_fill


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


# ── Sheet 1: Summary ────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"

ws.append([f"Scraping Progress Report - {today}"])
ws.merge_cells("A1:F1")
ws["A1"].font = Font(bold=True, size=14)
ws.append([])

summary_data = [
    ["Metric", "Value"],
    ["Total stores", total_stores],
    ["Stores complete", complete["count"]],
    ["Stores incomplete", incomplete["count"]],
    ["Stores pending", pending["count"]],
    ["Stores with no target", no_target],
    ["Reviews collected", total_scraped],
    ["Reviews in DB (incl. overshoot)", total_reviews],
    ["Target total (capped @1500)", total_target],
    ["Reviews remaining", total_remaining],
    ["Completion %", total_scraped / total_target if total_target else 0],
]
for i, row in enumerate(summary_data):
    ws.append(row)
    r = ws.max_row
    if i == 0:
        style_header(ws, r, 2)
    else:
        style_row(ws, r, 2)
ws.cell(row=ws.max_row, column=2).number_format = pct_fmt
for r in range(4, ws.max_row):
    ws.cell(row=r, column=2).number_format = num_fmt
auto_width(ws)

# ── Sheet 2: By Category ────────────────────────────────────────────────────
ws2 = wb.create_sheet("By Category")
headers = ["Category", "Stores", "Complete", "Incomplete", "Pending", "Collected", "Target", "Remaining", "% Done"]
ws2.append(headers)
style_header(ws2, 1, len(headers))

for r in major_cats:
    cat, cnt, done, inc, pend, collected, target, remain = r
    pct = collected / target if target else 0
    ws2.append([cat, cnt, done, inc, pend, collected, target, remain, pct])
    style_row(ws2, ws2.max_row, len(headers))
    ws2.cell(row=ws2.max_row, column=9).number_format = pct_fmt
    for c in range(2, 9):
        ws2.cell(row=ws2.max_row, column=c).number_format = num_fmt

# Total row
ws2.append(["TOTAL", t_cnt, t_c, t_i, t_p, t_col, t_tar, t_rem, t_col/t_tar if t_tar else 0])
style_row(ws2, ws2.max_row, len(headers), is_total=True)
ws2.cell(row=ws2.max_row, column=9).number_format = pct_fmt
for c in range(2, 9):
    ws2.cell(row=ws2.max_row, column=c).number_format = num_fmt
auto_width(ws2)

# ── Sheet 3: By Status ──────────────────────────────────────────────────────
ws3 = wb.create_sheet("By Status")
headers = ["Status", "Stores", "Target", "Collected", "Remaining", "% Done"]
ws3.append(headers)
style_header(ws3, 1, len(headers))
for name, d in [("complete", complete), ("incomplete", incomplete), ("pending", pending)]:
    rem = max(0, d["target"] - d["scraped"])
    pct = d["scraped"] / d["target"] if d["target"] else 0
    ws3.append([name, d["count"], d["target"], d["scraped"], rem, pct])
    style_row(ws3, ws3.max_row, len(headers))
    ws3.cell(row=ws3.max_row, column=6).number_format = pct_fmt
    for c in range(2, 6):
        ws3.cell(row=ws3.max_row, column=c).number_format = num_fmt
auto_width(ws3)

# ── Sheet 4: Batches ────────────────────────────────────────────────────────
if batch_info:
    ws4 = wb.create_sheet("Batches")
    headers = ["Batch", "Stores", "Target Low", "Target High", "Already Have", "To Scrape"]
    ws4.append(headers)
    style_header(ws4, 1, len(headers))
    for b in batch_info:
        ws4.append([b["name"], b["stores"], b["lo"], b["hi"], b["already"], b["delta"]])
        style_row(ws4, ws4.max_row, len(headers))
        for c in range(2, 7):
            ws4.cell(row=ws4.max_row, column=c).number_format = num_fmt
    # Total
    ws4.append(["TOTAL", sum(b["stores"] for b in batch_info), "", "",
                sum(b["already"] for b in batch_info), sum(b["delta"] for b in batch_info)])
    style_row(ws4, ws4.max_row, len(headers), is_total=True)
    for c in range(2, 7):
        ws4.cell(row=ws4.max_row, column=c).number_format = num_fmt
    auto_width(ws4)

# ── Sheet 5: Flagged Mismatches ──────────────────────────────────────────────
if flagged:
    ws5 = wb.create_sheet("Flagged Mismatches")
    headers = ["Store ID", "Name", "Address", "Place ID", "Target", "Scraped", "Excess", "Category"]
    ws5.append(headers)
    style_header(ws5, 1, len(headers))
    for r in flagged:
        ws5.append([r[0], r[1], r[3], r[2], r[4], r[5], r[6], r[7]])
        style_row(ws5, ws5.max_row, len(headers))
        for c in [5, 6, 7]:
            ws5.cell(row=ws5.max_row, column=c).number_format = num_fmt
    auto_width(ws5)

wb.save(XLSX_PATH)
print(f"Wrote {XLSX_PATH}")
print("Done!")
