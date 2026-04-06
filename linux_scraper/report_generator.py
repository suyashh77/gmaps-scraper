"""
Excel report generator — writes progress reports every N hours.

Generates .xlsx files with three sheets:
  1. Summary — overall stats
  2. Store Detail — per-store status breakdown
  3. Hourly Log — timestamped progress snapshots
"""

from __future__ import annotations

import logging
import os
import sqlite3
from datetime import datetime
from typing import Any

logger = logging.getLogger("report")


def _row_factory(cursor: sqlite3.Cursor, row: tuple) -> dict:
    cols = [desc[0] for desc in cursor.description]
    return dict(zip(cols, row))


def _safe_cell(val: Any) -> Any:
    """Prevent Excel formula injection by prefixing strings that start with =/@/+/-."""
    if isinstance(val, str) and val and val[0] in ('=', '@', '+', '-'):
        return "'" + val  # force Excel to treat as text
    return val


def _fmt_duration(seconds: float | None) -> str:
    if not seconds or seconds <= 0:
        return "N/A"
    hours, rem = divmod(int(seconds), 3600)
    minutes, sec = divmod(rem, 60)
    if hours >= 24:
        days, hours = divmod(hours, 24)
        return f"{days}d {hours}h {minutes}m"
    if hours:
        return f"{hours}h {minutes}m {sec}s"
    if minutes:
        return f"{minutes}m {sec}s"
    return f"{sec}s"


def generate_report(db_path: str, output_dir: str | None = None) -> str:
    """
    Generate an Excel progress report from the current DB state.

    Returns the path to the generated .xlsx file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = sqlite3.connect(db_path, timeout=10)
    conn.row_factory = _row_factory
    cur = conn.cursor()

    # ── Gather stats ─────────────────────────────────────────────────────
    cur.execute("SELECT COUNT(*) AS cnt FROM stores")
    total_stores = cur.fetchone()["cnt"]

    cur.execute("SELECT status, COUNT(*) AS cnt FROM stores GROUP BY status")
    by_status = {r["status"]: r["cnt"] for r in cur.fetchall()}

    cur.execute("SELECT COUNT(*) AS cnt FROM reviews")
    total_reviews = cur.fetchone()["cnt"]

    cur.execute(
        "SELECT AVG(reviews_scraped) AS avg_rev FROM stores WHERE status='completed'"
    )
    row = cur.fetchone()
    avg_reviews = row["avg_rev"] if row and row["avg_rev"] else 0

    cur.execute(
        "SELECT MIN(created_at) AS first_ts FROM stores"
    )
    row = cur.fetchone()
    first_ts = row["first_ts"] if row else None

    # Session info
    cur.execute("""
        SELECT SUM(stores_processed) AS sp,
               SUM(
                   CAST(strftime('%s', ended_at) AS INTEGER) -
                   CAST(strftime('%s', started_at) AS INTEGER)
               ) AS secs
        FROM scrape_sessions
        WHERE status IN ('completed', 'interrupted')
          AND ended_at IS NOT NULL AND started_at IS NOT NULL
    """)
    th = cur.fetchone()
    total_secs = th["secs"] or 0
    stores_processed_sessions = th["sp"] or 0

    # Per-store detail
    cur.execute("""
        SELECT store_id, input_name, input_city, input_state,
               status, reviews_scraped, gmaps_reviews_count,
               attempts, error_message, updated_at
        FROM stores
        ORDER BY store_id
    """)
    store_rows = cur.fetchall()

    # Session log
    cur.execute("""
        SELECT session_id, started_at, ended_at,
               stores_processed, stores_completed, stores_failed,
               reviews_collected, status
        FROM scrape_sessions
        ORDER BY session_id
    """)
    session_rows = cur.fetchall()

    conn.close()

    # ── Build workbook ───────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    number_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    completed = by_status.get("completed", 0)
    pending = by_status.get("pending", 0)
    in_progress = by_status.get("in_progress", 0)
    failed = by_status.get("failed", 0)
    skipped = by_status.get("skipped", 0)

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    ws.merge_cells("A1:B1")
    ws["A1"] = "Google Maps Scraper — Progress Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    data = [
        ("", ""),
        ("STORES", ""),
        ("Total Stores", total_stores),
        ("Completed", completed),
        ("Pending", pending),
        ("In Progress", in_progress),
        ("Failed", failed),
        ("Skipped (closed)", skipped),
        ("Completion %", f"{completed / total_stores * 100:.1f}%" if total_stores else "0%"),
        ("", ""),
        ("REVIEWS", ""),
        ("Total Reviews Collected", f"{total_reviews:,}"),
        ("Avg Reviews / Store", f"{avg_reviews:.0f}"),
        ("", ""),
        ("TIMING", ""),
        ("Total Scraping Time", _fmt_duration(total_secs)),
        ("Avg Time / Store", _fmt_duration(total_secs / stores_processed_sessions if stores_processed_sessions else 0)),
    ]

    for i, (label, value) in enumerate(data, start=4):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        if label and not value:
            ws[f"A{i}"].font = Font(bold=True, size=12, color="2F5496")
        elif label:
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"].font = number_font

    # ── Sheet 2: Store Detail ────────────────────────────────────────────
    ws2 = wb.create_sheet("Store Detail")
    headers = [
        "Store ID", "Name", "City", "State", "Status",
        "Reviews Scraped", "Google Review Count", "Attempts",
        "Error", "Last Updated",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, store in enumerate(store_rows, start=2):
        values = [
            store["store_id"],
            _safe_cell(store["input_name"] or ""),
            _safe_cell(store["input_city"] or ""),
            _safe_cell(store["input_state"] or ""),
            store["status"] or "",
            store["reviews_scraped"] or 0,
            store["gmaps_reviews_count"] or 0,
            store["attempts"] or 0,
            _safe_cell((store["error_message"] or "")[:80]),
            store["updated_at"] or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            # Color-code status
            if col_idx == 5:
                if val == "completed":
                    cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
                elif val == "failed":
                    cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                elif val == "pending":
                    cell.fill = PatternFill(start_color="FFEB9C", fill_type="solid")

    # Auto-fit columns
    for col_idx in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = max(
            len(headers[col_idx - 1]) + 4, 12
        )
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["I"].width = 50

    # ── Sheet 3: Session Log ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Session Log")
    s_headers = [
        "Session", "Started", "Ended", "Duration",
        "Stores Processed", "Completed", "Failed",
        "Reviews Collected", "Status",
    ]

    for col_idx, header in enumerate(s_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, sess in enumerate(session_rows, start=2):
        duration = ""
        if sess["started_at"] and sess["ended_at"]:
            try:
                s = datetime.fromisoformat(sess["started_at"])
                e = datetime.fromisoformat(sess["ended_at"])
                duration = _fmt_duration((e - s).total_seconds())
            except Exception:
                pass

        values = [
            sess["session_id"],
            sess["started_at"] or "",
            sess["ended_at"] or "",
            duration,
            sess["stores_processed"] or 0,
            sess["stores_completed"] or 0,
            sess["stores_failed"] or 0,
            sess["reviews_collected"] or 0,
            sess["status"] or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col_idx in range(1, len(s_headers) + 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = max(
            len(s_headers[col_idx - 1]) + 4, 14
        )

    # ── Save ─────────────────────────────────────────────────────────────
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(db_path), "reports")
    try:
        os.makedirs(output_dir, exist_ok=True)
    except OSError as e:
        # Fall back to the DB's directory if reports/ can't be created
        logger.warning(f"Could not create reports dir ({e}), saving alongside DB")
        output_dir = os.path.dirname(db_path) or "."

    timestamp = datetime.now().strftime("%Y-%m-%d_%Hh")
    filename = f"scraper_report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    try:
        wb.save(filepath)
    except PermissionError:
        # File may be open in Excel — save with a unique suffix instead
        import uuid
        filename = f"scraper_report_{timestamp}_{uuid.uuid4().hex[:4]}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        logger.warning(f"Original file was locked — saved as: {filepath}")
    logger.info(f"Report saved: {filepath}")
    return filepath
