"""
Update store targets from Excel and mark stores as complete or pending-rescrape.

Matching: place_id first, then name+address fuzzy fallback.
Target: reviews_outscraper column from the Excel.
Complete: (master_reviews + reviews_scraped) >= reviews_outscraper.
Rescrape: total_toward_target < reviews_outscraper -> reset to pending with attempts=0.
"""

import argparse
import re
import sqlite3
import sys

import openpyxl

EXCEL = "list2_forrester_outscrpenrich.xlsx"
DB = "list2_forrester_1000_reviews_a0ul1w.db"


def normalize(s: str) -> str:
    """Lowercase, strip, collapse whitespace, remove punctuation."""
    if not s:
        return ""
    s = re.sub(r"[^\w\s]", "", s.lower().strip())
    return re.sub(r"\s+", " ", s)


def load_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Update store targets from Excel and mark stores as complete or pending-rescrape"
    )
    parser.add_argument("--excel", "-e", default=EXCEL, help="Path to Excel file with targets")
    parser.add_argument("--db", "-d", default=DB, help="Path to SQLite database")
    args = parser.parse_args(argv)

    excel_rows = load_excel(args.excel)
    print(f"Loaded {len(excel_rows)} rows from Excel: {args.excel}")

    conn = sqlite3.connect(args.db)
    conn.row_factory = lambda c, r: dict(zip([d[0] for d in c.description], r))
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(stores)")
    store_cols = {r["name"] for r in cur.fetchall()}
    has_master_reviews = "master_reviews" in store_cols
    master_expr = "master_reviews" if has_master_reviews else "0 AS master_reviews"

    cur.execute(
        "SELECT store_id, google_place_id, input_name, gmaps_name, "
        "gmaps_address, input_street, input_city, input_state, "
        f"reviews_scraped, {master_expr}, target_reviews, status, attempts FROM stores"
    )
    db_stores = cur.fetchall()
    stores_by_id = {s["store_id"]: s for s in db_stores}

    print(f"Loaded {len(db_stores)} stores from DB")
    print(f"master_reviews column present: {has_master_reviews}")

    # Build lookup by place_id
    pid_lookup = {}
    for s in db_stores:
        pid = s["google_place_id"]
        if pid and pid.startswith("ChIJ"):
            pid_lookup[pid] = s

    # Build lookup by normalized name+address for fallback
    name_addr_lookup = {}
    for s in db_stores:
        name = normalize(s["gmaps_name"] or s["input_name"] or "")
        addr = normalize(s["gmaps_address"] or "")
        if not addr:
            parts = [s["input_street"] or "", s["input_city"] or "", s["input_state"] or ""]
            addr = normalize(" ".join(p for p in parts if p))
        if name:
            name_addr_lookup[(name, addr)] = s

    matched_rows = 0
    duplicate_store_matches = 0
    target_by_store: dict[int, int] = {}
    marked_complete = 0
    marked_rescrape = 0
    already_complete = 0
    unmatched = []

    for ex in excel_rows:
        target = ex.get("reviews_outscraper")
        if target is None:
            continue
        try:
            target = int(target)
        except (TypeError, ValueError):
            continue

        ex_pid = ex.get("place_id") or ""
        ex_name = normalize(ex.get("name") or "")
        ex_addr = normalize(ex.get("address") or "")

        # Match by place_id first
        store = None
        if ex_pid and ex_pid.startswith("ChIJ"):
            store = pid_lookup.get(ex_pid)

        # Fallback: name + address
        if not store and ex_name:
            store = name_addr_lookup.get((ex_name, ex_addr))
            if not store:
                # Try partial name match with minimum address overlap.
                for (db_name, db_addr), s in name_addr_lookup.items():
                    if ex_name and db_name and (ex_name in db_name or db_name in ex_name):
                        if ex_addr and db_addr:
                            ex_words = set(ex_addr.split())
                            db_words = set(db_addr.split())
                            if len(ex_words & db_words) >= 2:
                                store = s
                                break

        if not store:
            unmatched.append(
                f"  name={ex.get('name', '')[:40]}, addr={ex.get('address', '')[:50]}, "
                f"pid={ex_pid[:20]}, target={target}"
            )
            continue

        matched_rows += 1
        sid = int(store["store_id"])

        # If multiple Excel rows map to one store, keep the larger target
        # to avoid under-scraping.
        if sid in target_by_store:
            duplicate_store_matches += 1
            target_by_store[sid] = max(target_by_store[sid], target)
        else:
            target_by_store[sid] = target

    for sid, target in target_by_store.items():
        store = stores_by_id[sid]
        scraped = int(store.get("reviews_scraped") or 0)
        master = int(store.get("master_reviews") or 0)
        total_toward_target = scraped + master

        cur.execute(
            "UPDATE stores SET target_reviews=?, updated_at=datetime('now') WHERE store_id=?",
            (target, sid),
        )

        if total_toward_target >= target:
            if store["status"] == "completed":
                already_complete += 1
            else:
                cur.execute(
                    "UPDATE stores SET status='completed', worker_id=NULL, error_message=NULL, "
                    "updated_at=datetime('now') WHERE store_id=?",
                    (sid,),
                )
                marked_complete += 1
        else:
            cur.execute(
                "UPDATE stores SET status='pending', worker_id=NULL, attempts=0, "
                "error_message=NULL, updated_at=datetime('now') WHERE store_id=?",
                (sid,),
            )
            marked_rescrape += 1

    conn.commit()

    cur.execute("SELECT status, COUNT(*) as cnt FROM stores GROUP BY status")
    final_status = {r["status"]: r["cnt"] for r in cur.fetchall()}
    cur.execute("SELECT COUNT(*) as cnt FROM reviews")
    total_reviews = cur.fetchone()["cnt"]

    conn.close()

    print("\n=== Results ===")
    print(f"Matched rows:          {matched_rows}/{len(excel_rows)}")
    print(f"Unique stores matched: {len(target_by_store)}")
    print(f"Duplicate row matches: {duplicate_store_matches}")
    print(f"Already complete:      {already_complete}")
    print(f"Newly marked complete: {marked_complete}")
    print(f"Marked for rescrape:   {marked_rescrape}")
    print(f"Unmatched Excel rows:  {len(unmatched)}")

    if unmatched:
        print("\n=== Unmatched rows (first 20) ===")
        for u in unmatched[:20]:
            print(u)

    print("\n=== Final DB Status ===")
    for status, cnt in sorted(final_status.items()):
        print(f"  {status}: {cnt}")
    print(f"  Total reviews: {total_reviews:,}")


if __name__ == "__main__":
    sys.exit(main())
