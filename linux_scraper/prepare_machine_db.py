"""
Build a scraper-ready database from a store list + master reviews DB.

Takes an Excel/CSV file of stores and the master DB. Creates a new SQLite DB
in the scraper's schema with those stores + their existing reviews copied over
from master. The scraper's live dedup prevents re-collecting any review that
already exists.

Usage:
    python -m linux_scraper prepare --stores forrester.xlsx --master master.db
    python -m linux_scraper prepare --stores list_A.xlsx --master master.db --out machine_a.db

    # Pending stores (0 reviews) don't strictly need --master, but it won't hurt:
    python -m linux_scraper prepare --stores list_B.xlsx --master master.db
"""

import os
import sqlite3
import sys


def prepare(stores_path: str, master_path: str, out_path: str) -> None:
    if not os.path.exists(stores_path):
        print(f"ERROR: Store list not found: {stores_path}")
        sys.exit(1)
    if not os.path.exists(master_path):
        print(f"ERROR: Master DB not found: {master_path}")
        sys.exit(1)
    if os.path.abspath(master_path) == os.path.abspath(out_path):
        print("ERROR: --out cannot be the same file as --master")
        sys.exit(1)

    print(f"Stores: {stores_path}")
    print(f"Master: {master_path}")
    print(f"Output: {out_path}")
    print()

    # ── Read store list from Excel/CSV ───────────────────────────────────
    import openpyxl
    ext = os.path.splitext(stores_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(stores_path, read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        store_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            store_rows.append(dict(zip(headers, row)))
        wb.close()
    elif ext == ".csv":
        import csv
        with open(stores_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            store_rows = list(reader)
    else:
        print(f"ERROR: Unsupported format: {ext}")
        sys.exit(1)

    print(f"Stores in file: {len(store_rows)}")

    if not store_rows:
        print("No stores found!")
        return

    # ── Connect to master (for reviews lookup) ───────────────────────────
    master = sqlite3.connect(master_path, timeout=10)
    master.row_factory = lambda c, r: dict(zip([d[0] for d in c.description], r))
    mc = master.cursor()

    # Build master place_id → store_id lookup
    mc.execute("SELECT store_id, place_id FROM stores")
    master_pid_map = {}
    for r in mc.fetchall():
        if r["place_id"]:
            master_pid_map[r["place_id"]] = r["store_id"]

    # ── Create scraper DB ────────────────────────────────────────────────
    out_dir = os.path.dirname(os.path.abspath(out_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    if os.path.exists(out_path):
        if not os.path.isfile(out_path):
            print(f"ERROR: Output path exists and is not a file: {out_path}")
            sys.exit(1)
        os.remove(out_path)

    from linux_scraper.database import DatabaseManager
    db = DatabaseManager(out_path)
    db.connect()
    db.init_schema()
    cur = db.conn.cursor()

    # ── Insert stores ────────────────────────────────────────────────────
    incomplete_count = 0
    fresh_count = 0
    master_sid_to_scraper_sid = {}  # master store_id → scraper store_id

    for s in store_rows:
        master_reviews = 0
        rv = s.get("reviews_scraped")
        if rv is not None:
            try:
                master_reviews = int(rv)
            except (TypeError, ValueError):
                master_reviews = 0

        target = 0
        tv = s.get("target_reviews")
        if tv is not None:
            try:
                target = int(tv)
            except (TypeError, ValueError):
                target = 0

        if master_reviews > 0:
            incomplete_count += 1
        else:
            fresh_count += 1

        place_id = s.get("place_id") or ""

        cur.execute("""
            INSERT INTO stores (
                machine_id, input_name, input_street, input_city, input_state,
                input_zip, input_lat, input_lon, google_place_id,
                gmaps_name, gmaps_address, gmaps_phone, gmaps_website,
                gmaps_rating, gmaps_reviews_count, gmaps_lat, gmaps_lon,
                gmaps_category, gmaps_price_level,
                target_reviews, master_reviews, reviews_scraped,
                status
            ) VALUES (
                ?, ?, ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?,
                ?, ?, ?,
                ?
            )
        """, (
            "",
            s.get("store_name") or s.get("business_name") or "",
            s.get("street") or "",
            s.get("city") or "",
            s.get("state") or "NC",
            s.get("zip") or "",
            _float(s.get("source_lat")),
            _float(s.get("source_lon")),
            place_id,
            s.get("outscraper_name") or s.get("store_name") or "",
            s.get("full_address") or s.get("gmaps_address") or "",
            s.get("phone") or "",
            "",
            _float(s.get("rating")),
            target,
            None,
            None,
            s.get("category") or "",
            "",
            max(target, 1),
            master_reviews,
            0,          # reviews_scraped starts at 0 locally
            "pending",
        ))

        scraper_sid = cur.lastrowid

        # Map master store_id → scraper store_id for review copying
        master_sid = master_pid_map.get(place_id)
        if master_sid is not None:
            master_sid_to_scraper_sid[master_sid] = scraper_sid

    db.conn.commit()

    # ── Copy existing reviews from master ────────────────────────────────
    total_reviews_copied = 0
    stores_with_reviews = 0

    if master_sid_to_scraper_sid:
        print(f"Copying existing reviews for {len(master_sid_to_scraper_sid)} matched stores...")

        for master_sid, scraper_sid in master_sid_to_scraper_sid.items():
            mc.execute("""
                SELECT id, reviewer_name, rating, date_relative, review_text,
                       helpful_count, photo_count, has_owner_response,
                       reviewer_review_count, reviewer_photo_count, is_local_guide,
                       service_type, scraped_at
                FROM reviews
                WHERE store_id = ?
            """, (master_sid,))
            reviews = mc.fetchall()

            if not reviews:
                continue

            stores_with_reviews += 1
            rows = []
            for r in reviews:
                # Synthetic review_id: won't collide with real Google IDs
                review_id = f"m_{r['id']}"
                rows.append((
                    scraper_sid, "", review_id,
                    r.get("reviewer_name") or "Anonymous",
                    r.get("rating"),
                    r.get("date_relative") or "",
                    (r.get("review_text") or "")[:5000],
                    int(r.get("helpful_count") or 0),
                    int(r.get("photo_count") or 0),
                    int(r.get("has_owner_response") or 0),
                    int(r.get("reviewer_review_count") or 0),
                    int(r.get("reviewer_photo_count") or 0),
                    int(r.get("is_local_guide") or 0),
                    r.get("service_type") or "",
                    r.get("scraped_at") or "",
                ))

            cur.executemany("""
                INSERT OR IGNORE INTO reviews (
                    store_id, machine_id, review_id, reviewer_name, rating,
                    date_relative, review_text, helpful_count,
                    photo_count, has_owner_response, reviewer_review_count,
                    reviewer_photo_count, is_local_guide, service_type, scraped_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, rows)

            total_reviews_copied += len(rows)

            # Update local reviews_scraped count
            cur.execute(
                "UPDATE stores SET reviews_scraped=? WHERE store_id=?",
                (len(rows), scraper_sid),
            )

        db.conn.commit()

    # ── Summary ──────────────────────────────────────────────────────────
    cur.execute("SELECT COUNT(*) as cnt FROM stores")
    total_stores = cur.fetchone()["cnt"]
    cur.execute("SELECT COUNT(*) as cnt FROM reviews")
    total_reviews = cur.fetchone()["cnt"]
    cur.execute("SELECT SUM(target_reviews) as s FROM stores")
    total_target = cur.fetchone()["s"] or 0
    reviews_needed = total_target - total_reviews_copied

    db.close()
    master.close()

    print(f"\n{'='*50}")
    print(f"Scraper DB ready: {out_path}")
    print(f"{'='*50}")
    print(f"  Stores:            {total_stores}")
    print(f"    Incomplete:      {incomplete_count} (have reviews, need more)")
    print(f"    Fresh:           {fresh_count} (no reviews yet)")
    print(f"  Reviews copied:    {total_reviews_copied:,} ({stores_with_reviews} stores)")
    print(f"  Reviews needed:    {reviews_needed:,}")
    print(f"  DB size:           {os.path.getsize(out_path) / 1024 / 1024:.1f} MB")
    print()
    print(f"Run scraper:")
    print(f"  python -m linux_scraper --db {out_path}")


def _float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Build a scraper DB from a store list + master reviews",
    )
    parser.add_argument("--stores", required=True,
                        help="Excel/CSV file with stores to scrape")
    parser.add_argument("--master", required=True,
                        help="Path to master reviews DB (for existing reviews)")
    parser.add_argument("--out",
                        help="Output DB path (default: based on input filename)")
    args = parser.parse_args()

    if not args.out:
        base = os.path.splitext(os.path.basename(args.stores))[0]
        args.out = f"{base}_scraper.db"

    prepare(args.stores, args.master, args.out)


if __name__ == "__main__":
    main()
